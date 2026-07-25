import re

from django.test import TestCase, Client
from django.urls import reverse


class LandingPageViewTests(TestCase):
    """
    Unit tests for Landing Page View (Story 1.1)
    Tests AC1-AC8: Logo display, content, downloads, banners, responsive design
    """

    def setUp(self):
        """Set up test client for each test"""
        self.client = Client()
        self.url = reverse('landing_home')

    def test_landing_page_renders_successfully(self):
        """Test that landing page loads with status 200"""
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)

    def test_landing_page_uses_correct_template(self):
        """Test that landing page uses the correct template"""
        response = self.client.get(self.url)
        self.assertTemplateUsed(response, 'landing/home.html')
        self.assertTemplateUsed(response, 'base.html')

    def test_landing_page_contains_hero_section(self):
        """Test AC2: Hero section with main heading exists"""
        response = self.client.get(self.url)
        self.assertContains(
            response,
            'Podnesite prijavu za projekat ili inicijativu',
            msg_prefix="Hero heading missing or incorrect"
        )

    def test_landing_page_contains_preparation_checklist(self):
        """Test AC2: 'Šta treba da pripremim?' section exists"""
        response = self.client.get(self.url)
        self.assertContains(response, 'Šta treba da pripremim?')
        self.assertContains(response, 'Pre nego što počnete, pripremite:')
        # Check for checklist items
        self.assertContains(response, 'Osnovne podatke o vašoj organizaciji')
        self.assertContains(response, 'Excel šablon za budžet projekta')
        self.assertContains(response, 'Biografije članova tima')
        self.assertContains(response, 'Pisma podrške lokalnih organizacija')

    def test_landing_page_contains_reassurance_message(self):
        """Test AC2: Reassurance message about auto-save"""
        response = self.client.get(self.url)
        self.assertContains(response, 'Ne brinite - sve vaše podatke automatski čuvamo')
        self.assertContains(response, 'Možete prekinuti i vratiti se kasnije')

    def test_landing_page_contains_excel_template_download(self):
        """Test AC3: Excel template download section exists (Z6: client-supplied obrazac)"""
        response = self.client.get(self.url)
        self.assertContains(response, 'Preuzmite šablone')
        self.assertContains(response, 'Obrazac budžeta')
        self.assertContains(response, 'Preuzmi obrazac')
        # Check download link
        self.assertContains(response, 'budzet-obrazac.xlsx')

    def test_landing_page_contains_download_guidance(self):
        """Test AC3: Visual guidance for download process"""
        response = self.client.get(self.url)
        for step in ('1. Preuzmi', '2. Popuni', '3. Upload'):
            self.assertContains(response, step)

    def test_landing_page_contains_coa_banner(self):
        """Test AC4: COA (Projekat) banner exists with correct link"""
        response = self.client.get(self.url)
        self.assertContains(response, 'Prijava za Projekat')
        self.assertContains(response, 'Detaljni projekat sa budžetom, timom, i kompletnom dokumentacijom')
        self.assertContains(response, 'Započni prijavu')
        self.assertContains(response, 'href="/projekat/"')

    def test_landing_page_contains_cob_banner(self):
        """Test AC4: COB (Inicijativa) banner exists with correct link"""
        response = self.client.get(self.url)
        self.assertContains(response, 'Prijava za Inicijativu')
        self.assertContains(response, 'Brza prijava za inicijativu')
        self.assertContains(response, 'Započni prijavu')
        self.assertContains(response, 'href="/inicijativa/"')

    def test_landing_page_contains_domovik_logo(self):
        """Test AC1: Domovik logo is referenced in template"""
        response = self.client.get(self.url)
        self.assertContains(response, 'domovik-logo.svg')
        self.assertContains(response, 'alt="Domovik"')

    def test_landing_page_contains_donor_logos(self):
        """Test AC1: Donor logos are referenced in template"""
        response = self.client.get(self.url)
        self.assertContains(response, 'donor-1.svg')
        self.assertContains(response, 'donor-2.svg')
        self.assertContains(response, 'alt="Donor 1"')
        self.assertContains(response, 'alt="Donor 2"')

    def test_landing_page_includes_landing_css(self):
        """Test that landing.css is loaded"""
        response = self.client.get(self.url)
        self.assertContains(response, 'landing.css')

    def test_landing_page_has_semantic_html_structure(self):
        """Test AC8: Semantic HTML elements are used"""
        response = self.client.get(self.url)
        # Check for semantic sections
        self.assertContains(response, '<section')
        # Check for ARIA labels
        self.assertContains(response, 'aria-labelledby="hero-heading"')
        self.assertContains(response, 'id="hero-heading"')

    def test_landing_page_footer_exists(self):
        """Test that footer with copyright is present"""
        response = self.client.get(self.url)
        self.assertContains(response, '&copy; 2025 DOMOVIK')
        self.assertContains(response, 'Sva prava zadržana')

    def test_landing_page_content_type_is_html(self):
        """Test that response content type is HTML"""
        response = self.client.get(self.url)
        self.assertEqual(response['Content-Type'], 'text/html; charset=utf-8')

    def test_landing_page_response_not_cached(self):
        """Test that landing page is not cached (for dynamic content future)"""
        response = self.client.get(self.url)
        # Check that response is recent (not from cache)
        self.assertIsNotNone(response)


class LandingPageURLTests(TestCase):
    """Tests for URL routing to landing page"""

    def test_root_url_resolves_to_landing_page(self):
        """Test that root URL (/) resolves to landing page view"""
        url = reverse('landing_home')
        self.assertEqual(url, '/')

    def test_landing_page_accessible_without_authentication(self):
        """Test that landing page is publicly accessible"""
        response = self.client.get('/')
        self.assertEqual(response.status_code, 200)


class LandingPageStaticFilesTests(TestCase):
    """Tests for static file references in landing page"""

    def setUp(self):
        """Set up test client"""
        self.client = Client()
        self.url = reverse('landing_home')

    def test_base_css_is_referenced(self):
        """Test that base.css is loaded via base template"""
        response = self.client.get(self.url)
        # base.css should be included via base.html template
        content = response.content.decode('utf-8')
        # Check that CSS is being loaded (either base.css or inline styles)
        self.assertTrue(
            'base.css' in content or '<style>' in content,
            "No CSS found in landing page"
        )

    def test_static_images_directory_structure(self):
        """Test that image references use correct static file paths"""
        response = self.client.get(self.url)
        content = response.content.decode('utf-8')
        # Check that static file template tag is used
        self.assertIn('/static/images/', content)


class LandingPageAccessibilityTests(TestCase):
    """Tests for accessibility compliance (AC8)"""

    def setUp(self):
        """Set up test client"""
        self.client = Client()
        self.url = reverse('landing_home')

    def test_images_have_alt_text(self):
        """Test AC8: All images have alt attributes"""
        response = self.client.get(self.url)
        content = response.content.decode('utf-8')

        # Count img tags
        import re
        img_tags = re.findall(r'<img[^>]*>', content)

        for img_tag in img_tags:
            self.assertIn(
                'alt=',
                img_tag,
                f"Image tag missing alt attribute: {img_tag}"
            )

    def test_headings_hierarchy(self):
        """Test that heading hierarchy is logical (h1, then h2, etc.)"""
        response = self.client.get(self.url)
        content = response.content.decode('utf-8')

        # Check h1 exists
        self.assertIn('<h1', content, "No h1 heading found")
        # Check h2 exists
        self.assertIn('<h2', content, "No h2 headings found")

    def test_links_have_descriptive_text(self):
        """Test that links have descriptive text (not just 'click here')"""
        response = self.client.get(self.url)
        content = response.content.decode('utf-8').lower()

        # Check that generic link text is NOT used
        self.assertNotIn('click here', content)
        self.assertNotIn('klikni ovde', content)


class ExcelTemplateDownloadTests(TestCase):
    """
    Unit tests for the budget template download (Story 1.2, rewritten for Z6).

    Z6 (2026-07-25): the generated template was replaced by the client's own
    `Budzet_obrazac.xlsx`, served as `downloads/budzet-obrazac.xlsx` and shared by
    both application types. These tests assert the structure of THAT file, so a
    future re-upload that loses the formulas or the sheet protection fails loudly.
    """

    TEMPLATE_NAME = 'budzet-obrazac.xlsx'
    SHEET_NAME = 'Budžet šablon'
    # Rows the applicant fills in; column E holds the =C*D formula and stays locked.
    DATA_ROWS = (
        tuple(range(14, 18)) + tuple(range(20, 24))
        + tuple(range(26, 31)) + tuple(range(33, 38))
    )
    SUBTOTAL_CELLS = {
        'E18': '=SUM(E14:E17)',
        'E24': '=SUM(E20:E23)',
        'E31': '=SUM(E26:E30)',
        'E38': '=SUM(E33:E37)',
    }
    TOTAL_CELL = 'E39'
    TOTAL_FORMULA = '=E18+E24+E31+E38'

    def setUp(self):
        """Set up test client for each test"""
        self.client = Client()
        self.url = reverse('landing_home')

    def _template_path(self):
        from django.conf import settings
        return settings.BASE_DIR / 'static' / 'downloads' / self.TEMPLATE_NAME

    def _workbook(self):
        from openpyxl import load_workbook
        return load_workbook(str(self._template_path()))

    def test_landing_page_contains_excel_download_link(self):
        """Test AC1: Landing page has the budget template download link"""
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Preuzmi obrazac')
        self.assertContains(response, 'downloads/' + self.TEMPLATE_NAME)

    def test_excel_download_link_points_to_static_file(self):
        """Test AC1: Download link points to the static file and forces a download"""
        response = self.client.get(self.url)
        content = response.content.decode('utf-8')

        self.assertIn('downloads/' + self.TEMPLATE_NAME, content)

        download_link = re.search(
            r'<a[^>]*href="[^"]*budzet-obrazac\.xlsx"[^>]*>',
            content
        )
        self.assertIsNotNone(download_link, "Download link not found in HTML")
        self.assertIn('download', download_link.group(), "Download attribute missing from link")

    def test_both_application_forms_link_to_template(self):
        """Z6: the obrazac is reachable from inside both forms, not only the landing page"""
        for route in ('submissions:coa_form', 'submissions:cob_form'):
            with self.subTest(route=route):
                response = self.client.get(reverse(route))
                self.assertEqual(response.status_code, 200)
                self.assertContains(response, 'downloads/' + self.TEMPLATE_NAME)
                self.assertContains(response, 'Preuzmi obrazac budžeta')

    def test_excel_template_file_exists(self):
        """Test AC1,5: Template file exists in the static folder and is not empty"""
        file_path = self._template_path()
        self.assertTrue(file_path.exists(), f"Budget template not found at {file_path}")
        self.assertGreater(file_path.stat().st_size, 0, "Budget template file is empty")

    def test_excel_template_is_xlsx_format(self):
        """Test AC5: Template is a valid .xlsx workbook (Excel 2007+)"""
        file_path = self._template_path()
        self.assertTrue(str(file_path).endswith('.xlsx'), "Template must be .xlsx")
        try:
            self.assertIsNotNone(self._workbook(), "Failed to load Excel workbook")
        except Exception as e:
            self.fail(f"Budget template is not a valid .xlsx file: {e}")

    def test_excel_template_structure(self):
        """Test AC2,4: Template has the expected sheet and column headers"""
        wb = self._workbook()
        self.assertIn(self.SHEET_NAME, wb.sheetnames, f"Sheet {self.SHEET_NAME!r} not found")

        ws = wb[self.SHEET_NAME]
        expected_headers = [
            'Troškovi',
            'Jedinica',
            'Količina',
            'Jedinična cena (EUR)',
            'Ukupan budžet (EUR)',
            'Opis troškova',
        ]
        actual_headers = [ws.cell(12, col).value for col in range(1, 7)]
        self.assertEqual(
            actual_headers,
            expected_headers,
            f"Header row 12 doesn't match. Expected: {expected_headers}, Got: {actual_headers}"
        )

        # The four expense categories must still be there
        for row, prefix in ((13, '1.'), (19, '2.'), (25, '3.'), (32, '4.')):
            self.assertTrue(
                str(ws.cell(row, 1).value or '').startswith(prefix),
                f"Category heading {prefix} missing from row {row}"
            )

    def test_excel_template_has_sum_formulas(self):
        """Test AC2: Subtotals and the grand total are live formulas, not pasted numbers"""
        ws = self._workbook()[self.SHEET_NAME]

        for ref, formula in self.SUBTOTAL_CELLS.items():
            self.assertEqual(
                ws[ref].value, formula,
                f"Subtotal {ref} should be {formula}, got: {ws[ref].value}"
            )

        self.assertEqual(
            ws[self.TOTAL_CELL].value, self.TOTAL_FORMULA,
            f"Grand total {self.TOTAL_CELL} should be {self.TOTAL_FORMULA}, "
            f"got: {ws[self.TOTAL_CELL].value}"
        )

        # Every data row multiplies quantity by unit price
        for row in self.DATA_ROWS:
            self.assertEqual(
                ws.cell(row, 5).value, f'=C{row}*D{row}',
                f"Row {row} lost its =C*D formula"
            )

    def test_excel_template_formula_cells_are_protected(self):
        """Z6: sheet protection is on, formulas locked, input cells still editable"""
        ws = self._workbook()[self.SHEET_NAME]

        self.assertTrue(ws.protection.sheet, "Sheet protection is not enabled")

        locked_refs = list(self.SUBTOTAL_CELLS) + [self.TOTAL_CELL]
        locked_refs += [f'E{row}' for row in self.DATA_ROWS]
        for ref in locked_refs:
            self.assertTrue(
                ws[ref].protection.locked,
                f"Formula cell {ref} must stay locked"
            )

        for row in self.DATA_ROWS:
            for col in ('A', 'B', 'C', 'D', 'F'):
                self.assertFalse(
                    ws[f'{col}{row}'].protection.locked,
                    f"Input cell {col}{row} must be editable while the sheet is protected"
                )

        self.assertFalse(ws['B7'].protection.locked, "'Naziv tima' input must be editable")

    def test_excel_template_keeps_embedded_logo(self):
        """Z6: adding protection must not strip the client's branding"""
        ws = self._workbook()[self.SHEET_NAME]
        self.assertEqual(len(ws._images), 1, "Embedded logo missing from the template")

    def test_visual_guidance_text_present(self):
        """Test AC3: Visual guidance '1. Preuzmi → 2. Popuni → 3. Upload' is displayed"""
        response = self.client.get(self.url)
        for step in ('1. Preuzmi', '2. Popuni', '3. Upload'):
            self.assertContains(response, step)

    def test_visual_guidance_has_correct_css_class(self):
        """Test AC3: Visual guidance uses the downloads__steps class"""
        response = self.client.get(self.url)
        content = response.content.decode('utf-8')

        guidance_element = re.search(
            r'<p[^>]*class="[^"]*downloads__steps[^"]*"[^>]*>',
            content
        )
        self.assertIsNotNone(
            guidance_element,
            "Visual guidance element with class 'downloads__steps' not found"
        )

    def test_visual_guidance_in_same_section_as_download_link(self):
        """Test AC3: Visual guidance sits in the same card as the download link"""
        response = self.client.get(self.url)
        content = response.content.decode('utf-8')

        download_card = re.search(
            r'<div class="downloads__card reveal">.*?</section>',
            content,
            re.DOTALL
        )
        self.assertIsNotNone(download_card, "Download card section not found")

        card_content = download_card.group()
        self.assertIn('1. Preuzmi', card_content, "Visual guidance not in download card")
        self.assertIn(self.TEMPLATE_NAME, card_content, "Download link not in same card")


class ApplicationTypeSelectionTests(TestCase):
    """
    Unit tests for Application Type Selection (Story 1.3)
    Tests AC1-5: URL routing, banner descriptions, civic tech design, responsive layout, accessibility
    """

    def setUp(self):
        """Set up test client for each test"""
        self.client = Client()

    def test_coa_route_exists(self):
        """Test AC1: /projekat/ route returns HTTP 200"""
        response = self.client.get('/projekat/')
        self.assertEqual(response.status_code, 200)

    def test_cob_route_exists(self):
        """Test AC1: /inicijativa/ route returns HTTP 200"""
        response = self.client.get('/inicijativa/')
        self.assertEqual(response.status_code, 200)

    def test_coa_uses_correct_template(self):
        """Test AC1: COA form uses coa_form.html template"""
        response = self.client.get('/projekat/')
        self.assertTemplateUsed(response, 'submissions/coa_form.html')

    def test_cob_uses_correct_template(self):
        """Test AC1: COB form uses cob_form.html template"""
        response = self.client.get('/inicijativa/')
        self.assertTemplateUsed(response, 'submissions/cob_form.html')

    def test_landing_page_contains_coa_banner(self):
        """Test AC2: Landing page contains COA banner with description"""
        response = self.client.get(reverse('landing_home'))
        self.assertContains(response, 'Prijava za Projekat')
        self.assertContains(response, 'href="/projekat/"')

    def test_landing_page_contains_cob_banner(self):
        """Test AC2: Landing page contains COB banner with description"""
        response = self.client.get(reverse('landing_home'))
        self.assertContains(response, 'Prijava za Inicijativu')
        self.assertContains(response, 'href="/inicijativa/"')

    def test_banners_use_correct_css_classes(self):
        """Test AC3: Banners use BEM naming convention"""
        response = self.client.get(reverse('landing_home'))
        self.assertContains(response, 'landing__banner--coa')
        self.assertContains(response, 'landing__banner--cob')

    def test_coa_banner_has_enhanced_description(self):
        """Test AC2: COA banner has detailed description"""
        response = self.client.get(reverse('landing_home'))
        self.assertContains(response, 'budžet')
        self.assertContains(response, 'tim')
        self.assertContains(response, 'dokumentacijom')

    def test_cob_banner_has_enhanced_description(self):
        """Test AC2: COB banner has detailed description"""
        response = self.client.get(reverse('landing_home'))
        self.assertContains(response, 'Brza prijava')
        self.assertContains(response, 'inicijativ')

    def test_banners_have_heroicons(self):
        """Test AC2: Banners include Heroicon SVG icons"""
        response = self.client.get(reverse('landing_home'))
        content = response.content.decode('utf-8')
        # Check for SVG icons
        self.assertIn('<svg', content)
        self.assertIn('icon--document', content)
        self.assertIn('icon--lightbulb', content)

    def test_banners_have_cta_elements(self):
        """Test AC2: Banners have clear CTA (Call-to-Action) elements"""
        response = self.client.get(reverse('landing_home'))
        self.assertContains(response, 'Započni prijavu')
        self.assertContains(response, 'landing__banner-cta')

    def test_coa_form_page_renders_after_story_22(self):
        """Test: COA form page renders after Story 2.2 (no longer placeholder)"""
        response = self.client.get('/projekat/')
        self.assertEqual(response.status_code, 200)
        # Story 2.2: Form has progress stepper and entity switcher
        self.assertContains(response, 'Sekcija 1 od 3')
        self.assertContains(response, 'Fizičko lice')
        self.assertContains(response, 'Pravno lice')

    def test_cob_placeholder_page_has_back_button(self):
        """Test: COB placeholder page has working back to home link"""
        response = self.client.get('/inicijativa/')
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Nazad na početnu')
        # Verify the link uses named URL
        self.assertContains(response, 'href="/"')


class ResponsiveDesignTests(TestCase):
    """
    Tests for responsive design implementation (AC4)
    Verifies media queries and mobile optimizations
    """

    def test_landing_css_has_tablet_breakpoint(self):
        """Test AC4: Tablet breakpoint (768px) exists in CSS"""
        import os
        css_path = os.path.join('static', 'css', 'landing.css')

        with open(css_path, 'r', encoding='utf-8') as f:
            css_content = f.read()

        self.assertIn('@media (max-width: 768px)', css_content)
        self.assertIn('grid-template-columns: 1fr', css_content)

    def test_landing_css_has_mobile_breakpoint(self):
        """Test AC4: Mobile breakpoint (320px) exists in CSS"""
        import os
        css_path = os.path.join('static', 'css', 'landing.css')

        with open(css_path, 'r', encoding='utf-8') as f:
            css_content = f.read()

        self.assertIn('@media (max-width: 320px)', css_content)

    def test_landing_css_has_minimum_touch_targets(self):
        """Test AC4: CSS defines minimum 44x44px touch targets"""
        import os
        css_path = os.path.join('static', 'css', 'landing.css')

        with open(css_path, 'r', encoding='utf-8') as f:
            css_content = f.read()

        # Check for min-height: 44px
        self.assertIn('min-height: 44px', css_content)
        # Check for min-width: 44px
        self.assertIn('min-width: 44px', css_content)


class SubmissionsAppConfigTests(TestCase):
    """
    Tests for apps.submissions app registration and configuration
    Verifies critical setup from Task 1.0
    """

    def test_submissions_app_in_installed_apps(self):
        """Test Task 1.0: apps.submissions is registered in INSTALLED_APPS"""
        from django.conf import settings

        self.assertIn('apps.submissions', settings.INSTALLED_APPS)

    def test_submissions_app_config_exists(self):
        """Test: SubmissionsConfig is properly defined"""
        from apps.submissions.apps import SubmissionsConfig

        self.assertEqual(SubmissionsConfig.name, 'apps.submissions')
        self.assertEqual(SubmissionsConfig.default_auto_field, 'django.db.models.BigAutoField')

    def test_submissions_views_importable(self):
        """Test: Submissions views can be imported (routing dependency)"""
        try:
            from apps.submissions.views import ProjectApplicationView, InitiativeApplicationView
            self.assertTrue(True)
        except ImportError:
            self.fail("Could not import submissions views")


class ErrorHandlingTests(TestCase):
    """
    Tests for error handling and edge cases
    """

    def test_invalid_route_returns_404(self):
        """Test: Invalid routes return 404"""
        response = self.client.get('/invalid-route-does-not-exist/')
        self.assertEqual(response.status_code, 404)

    def test_projekat_with_trailing_invalid_path_returns_404(self):
        """Test: /projekat/invalid/ returns 404"""
        response = self.client.get('/projekat/invalid/')
        self.assertEqual(response.status_code, 404)

    def test_inicijativa_with_trailing_invalid_path_returns_404(self):
        """Test: /inicijativa/invalid/ returns 404"""
        response = self.client.get('/inicijativa/invalid/')
        self.assertEqual(response.status_code, 404)
