#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Script to verify and enhance Excel budget template for Story 1.2
Verifies structure, adds missing elements, and ensures formulas are correct

MEDIUM-4 Fix: Added UTF-8 encoding declaration for Windows console output
"""

import sys
import os
from copy import copy  # MEDIUM-2 Fix: Use copy() instead of deprecated .copy()
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.protection import SheetProtection

# MEDIUM-4 Fix: Configure stdout encoding for UTF-8 on Windows
if sys.stdout.encoding != 'utf-8':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except AttributeError:
        # Python < 3.7 fallback
        pass

EXCEL_PATH = "static/downloads/budzet-projekta-sablon.xlsx"

def verify_and_enhance_excel_template():
    """Verify and enhance Excel template structure"""

    print(f"[*] Loading Excel template: {EXCEL_PATH}")

    # Check if file exists
    # LOW-1 Fix: Added actionable error message with instructions
    if not os.path.exists(EXCEL_PATH):
        print(f"[ERROR] Excel file not found at {EXCEL_PATH}")
        print("[FIX] Create base Excel file first or restore from backup.")
        print("      This script verifies/enhances existing files - it does NOT create new files.")
        return False

    # Load workbook
    wb = load_workbook(EXCEL_PATH)

    print(f"[OK] Loaded workbook with {len(wb.sheetnames)} sheet(s): {wb.sheetnames}")

    # Verify/enhance main sheet
    if "Budžet Projekta" in wb.sheetnames:
        ws_budget = wb["Budžet Projekta"]
        print("[OK] Found 'Budzet Projekta' sheet")
    else:
        # Main sheet might have different name, check active sheet
        ws_budget = wb.active
        ws_budget.title = "Budžet Projekta"
        print("[WARN] Renamed active sheet to 'Budzet Projekta'")

    # ========================================================================
    # SECTION 1: Verify Column Headers
    # LOW-2 Fix: Added section comments for better code organization
    # ========================================================================
    expected_headers = [
        "Kategorija Troškova",
        "Opis",
        "Jedinična Cena (RSD)",
        "Količina",
        "Ukupno (RSD)"
    ]

    print("\n[*] Verifying column headers...")
    for col_idx, expected in enumerate(expected_headers, start=1):
        actual = ws_budget.cell(1, col_idx).value
        if actual == expected:
            print(f"  [OK] Column {col_idx}: {expected}")
        else:
            print(f"  [WARN] Column {col_idx}: Expected '{expected}', found '{actual}' - fixing...")
            ws_budget.cell(1, col_idx, expected)
            # Style header
            cell = ws_budget.cell(1, col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # ========================================================================
    # SECTION 2: Verify Predefined Budget Categories
    # ========================================================================
    print("\n[*] Verifying predefined categories...")
    categories = [
        "Plate i naknade",
        "Materijalni troškovi",
        "Usluge"
    ]

    for row_idx, category in enumerate(categories, start=2):
        actual = ws_budget.cell(row_idx, 1).value
        if actual == category:
            print(f"  [OK] Row {row_idx}: {category}")
        else:
            print(f"  [WARN] Row {row_idx}: Expected '{category}', found '{actual}' - fixing...")
            ws_budget.cell(row_idx, 1, category)

        # Add placeholder for Opis column
        if not ws_budget.cell(row_idx, 2).value:
            ws_budget.cell(row_idx, 2, "(primer)")

        # Add formula for Ukupno column (=C*D)
        formula_cell = ws_budget.cell(row_idx, 5)
        expected_formula = f"=C{row_idx}*D{row_idx}"
        if formula_cell.value != expected_formula:
            print(f"  [WARN] Row {row_idx}: Adding formula {expected_formula}")
            formula_cell.value = expected_formula

    # ========================================================================
    # SECTION 3: Add Empty Rows for User Budget Entries (rows 5-8)
    # ========================================================================
    for row_idx in range(5, 9):
        # Add formula for Ukupno column
        formula_cell = ws_budget.cell(row_idx, 5)
        expected_formula = f"=C{row_idx}*D{row_idx}"
        if not formula_cell.value or not str(formula_cell.value).startswith("="):
            formula_cell.value = expected_formula

    # ========================================================================
    # SECTION 4: Verify UKUPNO Row with SUM Formula
    # ========================================================================
    print("\n[*] Verifying UKUPNO row...")
    ukupno_row = 9
    ws_budget.cell(ukupno_row, 1, "UKUPNO")
    ws_budget.cell(ukupno_row, 1).font = Font(bold=True)

    # Add SUM formula (sums E2:E8)
    sum_formula = "=SUM(E2:E8)"
    sum_cell = ws_budget.cell(ukupno_row, 5)
    if sum_cell.value != sum_formula:
        print(f"  [WARN] Adding SUM formula: {sum_formula}")
        sum_cell.value = sum_formula
        sum_cell.font = Font(bold=True)
    else:
        print(f"  [OK] SUM formula correct: {sum_formula}")

    # ========================================================================
    # SECTION 5: Format Currency Columns (RSD)
    # ========================================================================
    print("\n[*] Formatting currency columns...")
    for row in ws_budget.iter_rows(min_row=2, max_row=9, min_col=3, max_col=5):
        for cell in row:
            cell.number_format = '#,##0.00 "RSD"'

    # ========================================================================
    # SECTION 6: Protect Formula Cells (Column E)
    # MEDIUM-2 Fix: Use copy() function instead of deprecated .copy() method
    # ========================================================================
    print("\n[*] Protecting formula cells...")
    # First, unlock all cells
    for row in ws_budget.iter_rows():
        for cell in row:
            cell.protection = copy(cell.protection)
            cell.protection.locked = False

    # Then lock only formula cells (column E, rows 2-9)
    for row in ws_budget.iter_rows(min_row=2, max_row=9, min_col=5, max_col=5):
        for cell in row:
            cell.protection = copy(cell.protection)
            cell.protection.locked = True

    # Protect sheet (no password, just lock formulas)
    ws_budget.protection.sheet = True
    ws_budget.protection.enable()
    print("  [OK] Formula cells protected (Column E locked)")

    # ========================================================================
    # SECTION 7: Set Column Widths for Readability
    # ========================================================================
    ws_budget.column_dimensions['A'].width = 25
    ws_budget.column_dimensions['B'].width = 30
    ws_budget.column_dimensions['C'].width = 20
    ws_budget.column_dimensions['D'].width = 12
    ws_budget.column_dimensions['E'].width = 20

    # ========================================================================
    # SECTION 8: Verify/Create Instructions Sheet (Uputstva)
    # ========================================================================
    print("\n[*] Checking for 'Uputstva' sheet...")
    if "Uputstva" not in wb.sheetnames and "Uputstvo" not in wb.sheetnames:
        print("  [WARN] 'Uputstva' sheet missing - creating...")
        ws_instructions = wb.create_sheet("Uputstva")

        # Add instructions content
        instructions_text = """UPUTSTVA ZA POPUNJAVANJE BUDŽETA

1. KATEGORIJE TROŠKOVA
   Odaberite kategoriju koja najbolje opisuje trošak:
   - Plate i naknade: Zarade za tim članova, konsultante
   - Materijalni troškovi: Kancelarijski materijal, oprema, potrepštine
   - Usluge: Zakup prostora, profesionalne usluge, prevoz

2. KAKO POPUNITI
   - Kolona "Opis": Detaljno opišite trošak (npr. "Laptop za projektnog menadžera")
   - Kolona "Jedinična Cena": Unesite cenu po jedinici u RSD (bez zareza)
   - Kolona "Količina": Unesite broj jedinica (npr. 2 laptopa)
   - Kolona "Ukupno": AUTOMATSKI SE RAČUNA - ne menjajte formulu!

3. UKUPNO
   Red "UKUPNO" automatski sabira sve troškove. Ovo je vaš totalni budžet projekta.

4. UPLOAD
   Kada završite, sačuvajte fajl i upload-ujte ga u Sekciji III prijave.

NAPOMENA: Preporučujemo Microsoft Excel ili LibreOffice Calc za najbolje rezultate."""

        # Split into lines and add to sheet
        lines = instructions_text.split('\n')
        for idx, line in enumerate(lines, start=1):
            cell = ws_instructions.cell(idx, 1, line)
            if line and (line.startswith('1.') or line.startswith('2.') or line.startswith('3.') or line.startswith('4.') or line.startswith('UPUTSTVA') or line.startswith('NAPOMENA')):
                cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        ws_instructions.column_dimensions['A'].width = 100
        print("  [OK] 'Uputstva' sheet created with instructions")
    else:
        existing_sheet = "Uputstva" if "Uputstva" in wb.sheetnames else "Uputstvo"
        print(f"  [OK] '{existing_sheet}' sheet already exists")

    # ========================================================================
    # SECTION 9: Save Workbook and Print Summary
    # ========================================================================
    print(f"\n[*] Saving enhanced template to {EXCEL_PATH}...")
    wb.save(EXCEL_PATH)
    print("[SUCCESS] Excel template verification and enhancement complete!")

    # Print summary
    file_size = os.path.getsize(EXCEL_PATH)
    print(f"\n[SUMMARY]")
    print(f"  - File: {EXCEL_PATH}")
    print(f"  - Size: {file_size} bytes")
    print(f"  - Sheets: {len(wb.sheetnames)} ({', '.join(wb.sheetnames)})")
    print(f"  - Format: Excel 2007+ (.xlsx)")

    return True

if __name__ == "__main__":
    success = verify_and_enhance_excel_template()
    exit(0 if success else 1)
