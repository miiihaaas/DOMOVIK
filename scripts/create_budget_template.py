#!/usr/bin/env python
"""
Create Excel Budget Template for DOMOVIK Project Applications
Story: 1-1-landing-page-with-information-guidance
Task: 4.1 - Create Excel template za budžet projekta
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path

def create_budget_template():
    """Create Excel budget template with formulas and formatting."""

    # Create workbook with two sheets
    wb = Workbook()

    # Sheet 1: Budget Template
    ws_budget = wb.active
    ws_budget.title = "Budžet Projekta"

    # Sheet 2: Instructions
    ws_instructions = wb.create_sheet("Uputstvo")

    # === BUDGET SHEET ===

    # Header row styling
    header_fill = PatternFill(start_color="006AAF", end_color="006AAF", fill_type="solid")
    header_font = Font(name='Helvetica Neue', size=12, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Define headers
    headers = ["Kategorija Troškova", "Opis", "Jedinična Cena (EUR)", "Količina", "Ukupno (EUR)"]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws_budget.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Set column widths
    ws_budget.column_dimensions['A'].width = 25
    ws_budget.column_dimensions['B'].width = 40
    ws_budget.column_dimensions['C'].width = 20
    ws_budget.column_dimensions['D'].width = 12
    ws_budget.column_dimensions['E'].width = 20

    # Sample categories and data
    sample_data = [
        ("Plate i naknade", "Plata koordinatora projekta (6 meseci)", 50000, 6),
        ("Plate i naknade", "Honorar volontera", 10000, 10),
        ("Materijalni troškovi", "Kancelarijski materijal", 5000, 1),
        ("Materijalni troškovi", "Štampanje materijala", 15000, 1),
        ("Usluge", "Iznajmljivanje sale za radionicu", 8000, 3),
        ("Usluge", "Internet i komunikacije", 3000, 6),
        ("Ostalo", "Nepredviđeni troškovi (max 10%)", 0, 1),
    ]

    # Border style
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Write sample data
    for row_num, (category, description, unit_price, quantity) in enumerate(sample_data, 2):
        ws_budget.cell(row=row_num, column=1, value=category).border = thin_border
        ws_budget.cell(row=row_num, column=2, value=description).border = thin_border
        ws_budget.cell(row=row_num, column=3, value=unit_price).border = thin_border
        ws_budget.cell(row=row_num, column=4, value=quantity).border = thin_border

        # Formula for Ukupno (Unit Price * Quantity)
        formula_cell = ws_budget.cell(row=row_num, column=5)
        formula_cell.value = f"=C{row_num}*D{row_num}"
        formula_cell.border = thin_border
        formula_cell.number_format = '#,##0'

    # Add Total row
    total_row = len(sample_data) + 2
    total_cell = ws_budget.cell(row=total_row, column=1)
    total_cell.value = "UKUPNO"
    total_cell.font = Font(bold=True, size=12)
    total_cell.fill = PatternFill(start_color="C6451F", end_color="C6451F", fill_type="solid")
    total_cell.font = Font(bold=True, size=12, color="FFFFFF")

    # Total formula
    total_formula_cell = ws_budget.cell(row=total_row, column=5)
    total_formula_cell.value = f"=SUM(E2:E{total_row-1})"
    total_formula_cell.font = Font(bold=True, size=12, color="FFFFFF")
    total_formula_cell.fill = PatternFill(start_color="C6451F", end_color="C6451F", fill_type="solid")
    total_formula_cell.number_format = '#,##0'

    # Data validation for numeric fields (Unit Price and Quantity)
    dv_number = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1=0)
    dv_number.error = "Molimo unesite pozitivan broj"
    dv_number.errorTitle = "Nevalidna vrednost"
    ws_budget.add_data_validation(dv_number)
    dv_number.add(f"C2:D{total_row-1}")

    # === INSTRUCTIONS SHEET ===

    ws_instructions['A1'] = "UPUTSTVO ZA POPUNJAVANJE BUDŽETA PROJEKTA"
    ws_instructions['A1'].font = Font(bold=True, size=14, color="006AAF")

    instructions_text = [
        "",
        "Koraci za popunjavanje:",
        "1. Preuzmite Excel šablon sa DOMOVIK platforme",
        "2. Popunite tabelu sa troškovima vašeg projekta",
        "3. Sačuvajte fajl i upload-ujte nazad na platformu",
        "",
        "Kolone u tabeli:",
        "• Kategorija Troškova - Izaberite kategoriju (Plate, Materijalni troškovi, Usluge, Ostalo)",
        "• Opis - Detaljan opis troška (npr. 'Plata koordinatora projekta za 6 meseci')",
        "• Jedinična Cena (EUR) - Cena po jedinici u dinarima",
        "• Količina - Broj jedinica (npr. broj meseci, broj ljudi, itd.)",
        "• Ukupno (EUR) - Automatski izračunato (ne menjajte formulu!)",
        "",
        "VAŽNO:",
        "✓ Kolona 'Ukupno' se automatski računa - ne menjajte formule",
        "✓ Unesite sve troškove koji su relevantni za vaš projekat",
        "✓ Ukupan budžet se prikazuje na dnu tabele",
        "✓ Maksimalno 10% budžeta može biti u kategoriji 'Nepredviđeni troškovi'",
        "",
        "Primer popunjene stavke:",
        "Kategorija: Materijalni troškovi",
        "Opis: Štampanje 500 flajera za promociju projekta",
        "Jedinična Cena: 10 RSD",
        "Količina: 500",
        "Ukupno: 5,000 RSD (automatski izračunato)",
        "",
        "Za pomoć kontaktirajte: domovik@example.com"
    ]

    for row_num, text in enumerate(instructions_text, 2):
        ws_instructions.cell(row=row_num, column=1, value=text)

    # Set column width for instructions
    ws_instructions.column_dimensions['A'].width = 80

    # Note: Cell protection can be added later if needed
    # For now, formulas are visible and can be examined by users

    # Save the workbook
    output_path = Path(__file__).parent.parent / "static" / "downloads" / "budzet-projekta-sablon.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"[OK] Excel template created successfully: {output_path}")
    print(f"[OK] File size: {output_path.stat().st_size / 1024:.1f} KB")
    return output_path

if __name__ == "__main__":
    create_budget_template()
